import logging
import json
from pathlib import Path
from timecode import Timecode
from openpyxl import Workbook, load_workbook
from .settings import SETTINGS
from .consts import UNWEIGHTED_RESULTS, NAME


def serialize_timecode(tc):
    if not isinstance(tc, Timecode):
        return tc
    return tc.__str__()


def write_batch_results(file_name_analyzed, plausibility, sequences):
    """ Schreibt errechnete Ergebnisse in Sammeldateien (JSON/Excel),
    je eine für alle in diesem Stapelverarbeitungsprozess analysierten Dateien.

    Args:
        * file_name_analyzed: Path
        * plausibility: int
        * sequences: [dict,..]
    """

    logger = logging.getLogger(NAME)
    logger.info('Schreibe Ergebnis in Sammeldatei')
    logger.debug(f'Exportiere: {sequences}')

    to_dir = SETTINGS['dir_results']

    file_name_id = SETTINGS['run_id']

    if plausibility == UNWEIGHTED_RESULTS:
        file_name_results = f'unschluessig_{file_name_analyzed.stem}_{file_name_id}'
    else:
        file_name_results = f'ergebnisse_{file_name_id}'

    excel_file = Path(to_dir) / Path(f'{file_name_results}.xlsx')
    json_file = Path(to_dir) / Path(f'{file_name_results}.json')

    for item in sequences:
        item.update(
            dict(plausibility=plausibility,
                 in_file=str(file_name_analyzed.stem))
        )

    sequences_to_json(json_file, sequences)
    sequences_to_excel(excel_file, sequences)


def sequences_to_excel(excel_file, sequences):

    if Path(excel_file).is_file():
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
    ws = wb.active

    row = len(list(ws.rows))

    if row == 1:
        ws.cell(row=row, column=1, value='Untersuchte Datei')
        ws.cell(row=row, column=2, value='Korpusnummer')
        ws.cell(row=row, column=3, value='Typ')
        ws.cell(row=row, column=4, value='Wert')

    for sequence in sequences:
        for key in sequence:
            row += 1
            ws.cell(row=row, column=1, value=sequence['in_file'])
            ws.cell(row=row, column=2, value=sequence['sequence_nr'])
            ws.cell(row=row, column=3, value=key)
            ws.cell(row=row, column=4, value=serialize_timecode(sequence[key]))

    wb.save(excel_file)


def sequences_to_json(json_file, sequences):
    file_data = [sequences]
    if Path(json_file).is_file():
        with open(json_file, 'r') as f:
            file_data = json.load(f)
        file_data.append(sequences)
    with open(json_file, 'w') as f:
        json.dump(file_data, f, indent=4, default=serialize_timecode)


def write_single_results(file_name_analyzed, plausibility, sequences):
    """ Schreibt errechnete Ergebnisse in Einzeldateien (JSON/Excel)
    pro zu analysierende Datei.

    Args:
        * file_name_analyzed: Path
        * plausibility: int
        * sequences: [dict,..]
    """

    logger = logging.getLogger(NAME)
    logger.info('Schreibe Ergebnis in Einzeldateien')
    logger.debug(f'Exportiere: {sequences}')
    to_dir = SETTINGS['dir_results']

    file_name_id = SETTINGS['run_id']
    file_name_results = f'{Path(file_name_analyzed).stem}_{file_name_id}'

    excel_file = Path(to_dir) / Path(f'{file_name_results}.xlsx')
    json_file = Path(to_dir) / Path(f'{file_name_results}.json')

    for item in sequences:
        item.update(
            dict(plausibility=plausibility,
                 in_file=str(file_name_analyzed.stem)
                 ))

    sequences_to_json(json_file, sequences)
    sequences_to_excel(excel_file, sequences)


def write_results(file_name_analyzed, plausibility, sequences):
    """ Löst entweder Sammel- oder Einzeldateien-Export der Ergebnisse aus
    (abhängig von SETTINGS['batch_processing']).

    Args:
        * file_name_analyzed: Path
        * plausibility: int
        * sequences: [dict,..]
    """
    logger = logging.getLogger(NAME)
    logger.debug(
        f'input: {file_name_analyzed, plausibility, sequences}')

    if plausibility == UNWEIGHTED_RESULTS:
        write_batch_results(file_name_analyzed, plausibility, sequences)
        return None

    if SETTINGS['batch_processing']:
        write_batch_results(file_name_analyzed, plausibility, sequences)
        return None

    write_single_results(file_name_analyzed, plausibility, sequences)
